"""Excel model export — a live, formula-driven workbook, banker conventions.

Sheets:
  Cover        summary card (green cross-sheet links)
  Assumptions  ALL inputs in blue: market data, cost of capital, per-year drivers
  WACC         CAPM build, formulas off Assumptions
  Model        historical actuals (blue, sourced) + forecast columns (formulas)
  DCF          discounting, terminal value (g/ROIC-consistent), EV→equity bridge
  Sensitivity  5×5 live grid (WACC × g), every cell a self-contained formula
  IS / BS / CF reported statements as provided by the data source

Conventions (industry standard):
  blue  = hardcoded inputs the user can change
  black = formulas
  green = links pulling from another sheet
  yellow fill = key assumptions to review
  negatives in parentheses; zeros as "-"; years as text; money in millions.

Change any blue cell (growth, margin, beta, ERP...) and the whole model,
including the sensitivity grid, recalculates.
"""
from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schema import CompanyData

# ---- style kit -------------------------------------------------------------
F = "Arial"
BLUE = Font(name=F, size=10, color="0000FF")            # inputs
BLACK = Font(name=F, size=10, color="000000")           # formulas
GREEN = Font(name=F, size=10, color="008000")           # cross-sheet links
BOLD = Font(name=F, size=10, bold=True)
H1 = Font(name=F, size=14, bold=True)
H2 = Font(name=F, size=11, bold=True, color="0F6466")
MUT = Font(name=F, size=9, color="6B6F76")
YEL = PatternFill("solid", start_color="FFFF00")
HDR_FILL = PatternFill("solid", start_color="0E1116")
HDR_FONT = Font(name=F, size=10, bold=True, color="FFFFFF")
HIST_FILL = PatternFill("solid", start_color="F1F0EA")
TOPLINE = Border(top=Side(style="thin"))
DBL = Border(top=Side(style="double"))

MM = '#,##0;(#,##0);"-"'          # millions
PS = '#,##0.00;(#,##0.00);"-"'    # per share
PCT = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
MULT = '0.0"x"'


def _w(ws, r, c, v, font=BLACK, fmt=None, fill=None, bold=False, border=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name=font.name, size=font.size, color=font.color,
                     bold=bold or font.bold)
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if border: cell.border = border
    return cell


def _title(ws, text, sub=""):
    _w(ws, 1, 1, text, H1)
    if sub: _w(ws, 2, 1, sub, MUT)
    ws.sheet_view.showGridLines = False


def _hdr_row(ws, r, labels, start_col=1):
    for i, lab in enumerate(labels):
        c = ws.cell(row=r, column=start_col + i, value=lab)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="right" if i else "left")


# --------------------------------------------------------------------------- #
def export_model(data: CompanyData, res: dict, path: str) -> str:
    """Build the workbook from a value_company()/value_data() result dict."""
    dcf = res.get("dcf")
    bd = res.get("base_drivers")
    wacc = res["wacc"]
    n = bd.horizon if bd else 5
    y0 = data.latest.year
    cur = data.currency
    scale = 1e6

    wb = Workbook()

    # ======================= ASSUMPTIONS ==================================== #
    ws = wb.create_sheet("Assumptions")
    _title(ws, f"{data.ticker} — Assumptions & Drivers",
           "Blue cells are inputs — change them and the whole model recalculates. "
           f"Source: {data.provider}, retrieved {date.today().isoformat()}.")
    ws.column_dimensions["A"].width = 34
    for col in "BCDEFGHIJK": ws.column_dimensions[col].width = 12

    r = 4
    _w(ws, r, 1, "Market data", H2); r += 1
    _w(ws, r, 1, f"Share price ({cur})"); _w(ws, r, 2, data.price, BLUE, PS); PRICE = f"Assumptions!$B${r}"; r += 1
    _w(ws, r, 1, "Diluted shares (mn)"); _w(ws, r, 2, data.shares_diluted / scale, BLUE, MM); SHARES = f"Assumptions!$B${r}"; r += 1
    _w(ws, r, 1, f"Total debt ({cur}mm)"); _w(ws, r, 2, data.latest.total_debt / scale, BLUE, MM); DEBT = f"Assumptions!$B${r}"; r += 1
    _w(ws, r, 1, f"Cash & ST investments ({cur}mm)"); _w(ws, r, 2, data.latest.cash_and_sti / scale, BLUE, MM); CASH = f"Assumptions!$B${r}"; r += 1
    _w(ws, r, 1, f"Minority interest ({cur}mm)"); _w(ws, r, 2, data.latest.minority_interest / scale, BLUE, MM); MIN = f"Assumptions!$B${r}"; r += 1
    _w(ws, r, 1, f"Net debt ({cur}mm)")
    _w(ws, r, 2, f"={DEBT}-{CASH}", BLACK, MM); NETDEBT = f"Assumptions!$B${r}"; r += 2

    _w(ws, r, 1, "Cost of capital inputs", H2); r += 1
    def inp(label, val, fmt=PCT2, yellow=False):
        nonlocal r
        _w(ws, r, 1, label)
        c = _w(ws, r, 2, val, BLUE, fmt)
        if yellow: c.fill = YEL
        ref = f"Assumptions!$B${r}"; r += 1
        return ref
    RF = inp("Risk-free rate (Rf)", wacc.risk_free, yellow=True)
    ERP = inp("Equity risk premium (ERP)", wacc.erp, yellow=True)
    CTY = inp("Country risk premium", wacc.country_premium)
    SZP = inp("Size premium", wacc.size_premium)
    BETA = inp("Beta (levered)", round(wacc.beta_used, 2), '0.00', yellow=True)
    KD = inp("Pre-tax cost of debt", wacc.cost_of_debt_pretax)
    TAX = inp("Tax rate", bd.tax_rate if bd else wacc.tax_rate, PCT)
    r += 1

    _w(ws, r, 1, "Terminal assumptions", H2); r += 1
    TG = inp("Terminal growth (g)", bd.terminal_growth if bd else 0.025, yellow=True)
    r += 1

    DRV = {}
    if bd:
        _w(ws, r, 1, "Per-year forecast drivers", H2); r += 1
        _hdr_row(ws, r, ["Driver"] + [f"FY{y0+i+1}E" for i in range(n)]); hdr = r; r += 1
        rows = [("Revenue growth", bd.revenue_growth, PCT),
                ("EBIT margin", bd.ebit_margin, PCT),
                ("D&A % of revenue", bd.da_pct, PCT),
                ("Capex % of revenue", bd.capex_pct, PCT)]
        for lab, vals, fmt in rows:
            _w(ws, r, 1, lab)
            for i, v in enumerate(vals):
                _w(ws, r, 2 + i, v, BLUE, fmt)
            DRV[lab] = r; r += 1
        _w(ws, r, 1, "ΔNWC % of Δrevenue")
        _w(ws, r, 2, bd.nwc_pct_delta, BLUE, PCT); DRV["nwc"] = r; r += 1

    A = dict(PRICE=PRICE, SHARES=SHARES, DEBT=DEBT, CASH=CASH, MIN=MIN,
             NETDEBT=NETDEBT, RF=RF, ERP=ERP, CTY=CTY, SZP=SZP, BETA=BETA,
             KD=KD, TAX=TAX, TG=TG, DRV=DRV)

    # ============================ WACC ====================================== #
    ws = wb.create_sheet("WACC")
    _title(ws, "Cost of capital (CAPM)", "All formulas reference the Assumptions sheet.")
    ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 14
    r = 4
    _w(ws, r, 1, "Cost of equity (Ke = Rf + β·ERP + CRP + size)")
    _w(ws, r, 2, f"={A['RF']}+{A['BETA']}*{A['ERP']}+{A['CTY']}+{A['SZP']}", GREEN, PCT2); KE = f"WACC!$B${r}"; r += 1
    _w(ws, r, 1, "After-tax cost of debt (Kd·(1−t))")
    _w(ws, r, 2, f"={A['KD']}*(1-{A['TAX']})", GREEN, PCT2); KDA = f"WACC!$B${r}"; r += 1
    _w(ws, r, 1, "Market value of equity (mm)")
    _w(ws, r, 2, f"={A['PRICE']}*{A['SHARES']}", GREEN, MM); MVE = f"WACC!$B${r}"; r += 1
    _w(ws, r, 1, "Debt (book, mm)")
    _w(ws, r, 2, f"={A['DEBT']}", GREEN, MM); MVD = f"WACC!$B${r}"; r += 1
    _w(ws, r, 1, "Weight of equity"); _w(ws, r, 2, f"={MVE}/({MVE}+{MVD})", BLACK, PCT); WE = f"WACC!$B${r}"; r += 1
    _w(ws, r, 1, "Weight of debt"); _w(ws, r, 2, f"=1-{WE}", BLACK, PCT); WD = f"WACC!$B${r}"; r += 1
    _w(ws, r, 1, "WACC", BOLD)
    _w(ws, r, 2, f"={WE}*{KE}+{WD}*{KDA}", BLACK, PCT2, bold=True, border=DBL); WACCC = f"WACC!$B${r}"; r += 1

    # ============================ MODEL ===================================== #
    MODEL = {}
    if bd and dcf:
        ws = wb.create_sheet("Model")
        hist = data.years[-4:] if len(data.years) >= 4 else data.years
        nh = len(hist)
        _title(ws, f"Operating model ({cur}mm)",
               "Shaded columns are reported actuals (blue hardcodes, sourced); "
               "unshaded are forecast formulas driven by the Assumptions sheet.")
        ws.column_dimensions["A"].width = 30
        for i in range(nh + n): ws.column_dimensions[get_column_letter(2 + i)].width = 12
        r = 4
        labels = [f"FY{h.year}A" for h in hist] + [f"FY{y0+i+1}E" for i in range(n)]
        _hdr_row(ws, r, ["Line item"] + labels); r += 1
        first_fc = 2 + nh                                    # first forecast column
        def col(i): return get_column_letter(i)

        def hist_row(label, vals, fmt=MM):
            nonlocal r
            _w(ws, r, 1, label)
            for i, v in enumerate(vals):
                _w(ws, r, 2 + i, v / scale, BLUE, fmt, fill=HIST_FILL)
            row = r; r += 1
            return row

        # Revenue: hist blue, forecast = prev*(1+growth)
        REV = hist_row("Revenue", [h.revenue for h in hist])
        r -= 1
        for j in range(n):
            c = first_fc + j
            g_ref = f"Assumptions!{col(2+j)}${A['DRV']['Revenue growth']}"
            _w(ws, r, c, f"={col(c-1)}{REV}*(1+{g_ref})", BLACK, MM)
        r += 1
        _w(ws, r, 1, "  growth %")
        for i in range(1, nh + n):
            c = 2 + i
            _w(ws, r, c, f"={col(c)}{REV}/{col(c-1)}{REV}-1", BLACK, PCT,
               fill=HIST_FILL if i < nh else None)
        GRW = r; r += 1

        EBIT = hist_row("EBIT", [h.ebit for h in hist]); r -= 1
        for j in range(n):
            c = first_fc + j
            m_ref = f"Assumptions!{col(2+j)}${A['DRV']['EBIT margin']}"
            _w(ws, r, c, f"={col(c)}{REV}*{m_ref}", BLACK, MM)
        r += 1
        _w(ws, r, 1, "  EBIT margin %")
        for i in range(nh + n):
            c = 2 + i
            _w(ws, r, c, f"={col(c)}{EBIT}/{col(c)}{REV}", BLACK, PCT,
               fill=HIST_FILL if i < nh else None)
        r += 1

        DA = hist_row("D&A", [h.depreciation_amort for h in hist]); r -= 1
        for j in range(n):
            c = first_fc + j
            _w(ws, r, c, f"={col(c)}{REV}*Assumptions!{col(2+j)}${A['DRV']['D&A % of revenue']}", BLACK, MM)
        r += 1
        _w(ws, r, 1, "EBITDA")
        for i in range(nh + n):
            c = 2 + i
            _w(ws, r, c, f"={col(c)}{EBIT}+{col(c)}{DA}", BLACK, MM,
               fill=HIST_FILL if i < nh else None)
        EBITDA = r; r += 1

        _w(ws, r, 1, "NOPAT (EBIT×(1−t))")
        for i in range(nh + n):
            c = 2 + i
            _w(ws, r, c, f"={col(c)}{EBIT}*(1-{A['TAX']})", GREEN, MM,
               fill=HIST_FILL if i < nh else None)
        NOPAT = r; r += 1

        CAPEX = hist_row("Capex", [h.capex for h in hist]); r -= 1
        for j in range(n):
            c = first_fc + j
            _w(ws, r, c, f"={col(c)}{REV}*Assumptions!{col(2+j)}${A['DRV']['Capex % of revenue']}", BLACK, MM)
        r += 1

        NWC = hist_row("Δ Net working capital", [h.change_in_nwc for h in hist]); r -= 1
        for j in range(n):
            c = first_fc + j
            _w(ws, r, c, f"=({col(c)}{REV}-{col(c-1)}{REV})*Assumptions!$B${A['DRV']['nwc']}", BLACK, MM)
        r += 1

        _w(ws, r, 1, "FCFF", BOLD)
        for i in range(nh + n):
            c = 2 + i
            _w(ws, r, c, f"={col(c)}{NOPAT}+{col(c)}{DA}-{col(c)}{CAPEX}-{col(c)}{NWC}",
               BLACK, MM, bold=True, border=TOPLINE, fill=HIST_FILL if i < nh else None)
        FCFF = r; r += 1
        MODEL = dict(REV=REV, EBIT=EBIT, DA=DA, EBITDA=EBITDA, NOPAT=NOPAT,
                     CAPEX=CAPEX, NWC=NWC, FCFF=FCFF, first_fc=first_fc, nh=nh)

    # ============================= DCF ====================================== #
    if bd and dcf and MODEL:
        ws = wb.create_sheet("DCF")
        _title(ws, f"Discounted cash flow ({cur}mm)",
               "Mid-year convention. Terminal value is reinvestment-consistent: "
               "TV = NOPATn·(1+g)·(1−g/ROIC)/(WACC−g).")
        ws.column_dimensions["A"].width = 38
        for i in range(n + 1): ws.column_dimensions[get_column_letter(2 + i)].width = 12
        fc0 = MODEL["first_fc"]
        def mcol(j): return get_column_letter(fc0 + j)      # model forecast col j
        def col(i): return get_column_letter(i)

        r = 4
        _hdr_row(ws, r, ["Year"] + [f"FY{y0+i+1}E" for i in range(n)]); r += 1
        _w(ws, r, 1, "FCFF (from Model)")
        for j in range(n):
            _w(ws, r, 2 + j, f"=Model!{mcol(j)}{MODEL['FCFF']}", GREEN, MM)
        FC = r; r += 1
        _w(ws, r, 1, "Period (mid-year, t−0.5)")
        for j in range(n):
            _w(ws, r, 2 + j, j + 0.5, BLACK, '0.0')
        T = r; r += 1
        _w(ws, r, 1, "Discount factor")
        for j in range(n):
            c = 2 + j
            _w(ws, r, c, f"=1/(1+{WACCC})^{col(c)}{T}", GREEN, '0.000')
        DF = r; r += 1
        _w(ws, r, 1, "PV of FCFF")
        for j in range(n):
            c = 2 + j
            _w(ws, r, c, f"={col(c)}{FC}*{col(c)}{DF}", BLACK, MM)
        PV = r; r += 2

        B = "B"
        _w(ws, r, 1, "PV of explicit FCFF")
        _w(ws, r, 2, f"=SUM(B{PV}:{col(1+n)}{PV})", BLACK, MM); PVE = f"$B${r}"; r += 1
        lastm = mcol(n - 1)
        _w(ws, r, 1, "Terminal NOPAT (FYn)")
        _w(ws, r, 2, f"=Model!{lastm}{MODEL['NOPAT']}", GREEN, MM); NOPn = f"$B${r}"; r += 1
        _w(ws, r, 1, "Invested capital (mm)")
        _w(ws, r, 2, (data.latest.invested_capital or
                      (data.latest.total_debt + data.latest.total_equity
                       - data.latest.cash_and_sti)) / scale, BLUE, MM)
        IC0 = f"$B${r}"; r += 1
        _w(ws, r, 1, "Terminal invested capital (scaled w/ revenue)")
        _w(ws, r, 2, f"={IC0}*Model!{lastm}{MODEL['REV']}/Model!{get_column_letter(1+MODEL['nh'])}{MODEL['REV']}",
           GREEN, MM); ICn = f"$B${r}"; r += 1
        _w(ws, r, 1, "Terminal ROIC")
        _w(ws, r, 2, f"=MAX({NOPn}/{ICn},{WACCC})", BLACK, PCT); ROIC = f"$B${r}"; r += 1
        _w(ws, r, 1, "Terminal reinvestment rate (g/ROIC)")
        _w(ws, r, 2, f"={A['TG']}/{ROIC}", GREEN, PCT); RRT = f"$B${r}"; r += 1
        _w(ws, r, 1, "Terminal value (Gordon)")
        _w(ws, r, 2, f"={NOPn}*(1+{A['TG']})*(1-{RRT})/({WACCC}-{A['TG']})",
           BLACK, MM); TV = f"$B${r}"; r += 1
        _w(ws, r, 1, "PV of terminal value")
        _w(ws, r, 2, f"={TV}/(1+{WACCC})^({n}-0.5)", BLACK, MM); PVTV = f"$B${r}"; r += 1
        _w(ws, r, 1, "TV as % of EV")
        _w(ws, r, 2, f"={PVTV}/({PVE}+{PVTV})", BLACK, PCT); r += 2

        _w(ws, r, 1, "Enterprise value", BOLD)
        _w(ws, r, 2, f"={PVE}+{PVTV}", BLACK, MM, bold=True, border=TOPLINE); EV = f"$B${r}"; r += 1
        _w(ws, r, 1, "(−) Net debt"); _w(ws, r, 2, f"=-{A['NETDEBT']}", GREEN, MM); r += 1
        _w(ws, r, 1, "(−) Minority interest"); _w(ws, r, 2, f"=-{A['MIN']}", GREEN, MM); r += 1
        _w(ws, r, 1, "Equity value", BOLD)
        _w(ws, r, 2, f"={EV}-{A['NETDEBT']}-{A['MIN']}", BLACK, MM, bold=True, border=TOPLINE); EQ = f"$B${r}"; r += 1
        _w(ws, r, 1, "Value per share", BOLD)
        _w(ws, r, 2, f"={EQ}/{A['SHARES']}", BLACK, PS, bold=True); VPS = f"$B${r}"; VPSROW = r; r += 1
        _w(ws, r, 1, "Upside vs price", BOLD)
        _w(ws, r, 2, f"={VPS}/{A['PRICE']}-1", GREEN, PCT, bold=True, border=DBL); r += 1
        DCFREF = dict(FC=FC, T=T, n=n, NOPn=NOPn, ROIC=ROIC, VPSROW=VPSROW)

        # ========================= SENSITIVITY ============================== #
        ws = wb.create_sheet("Sensitivity")
        _title(ws, f"Value per share — WACC × terminal growth ({cur}/share)",
               "Every cell is a live formula; it re-runs the full DCF at that WACC/g pair.")
        ws.column_dimensions["A"].width = 12
        base_w = wacc.wacc
        base_g = bd.terminal_growth
        waccs = [base_w + d for d in (-0.015, -0.0075, 0, 0.0075, 0.015)]
        gs = [base_g + d for d in (-0.01, -0.005, 0, 0.005, 0.01)]
        gs = [g for g in gs if g < min(waccs) - 0.005] or [base_g]
        r0 = 4
        _w(ws, r0, 1, "WACC \\ g", BOLD)
        for j, g in enumerate(gs):
            _w(ws, r0, 2 + j, g, BLUE, PCT)
            ws.column_dimensions[get_column_letter(2 + j)].width = 11
        fcff_rng = f"DCF!$B${DCFREF['FC']}:${get_column_letter(1+n)}${DCFREF['FC']}"
        t_rng = f"DCF!$B${DCFREF['T']}:${get_column_letter(1+n)}${DCFREF['T']}"
        for i, w_ in enumerate(waccs):
            rr = r0 + 1 + i
            _w(ws, rr, 1, w_, BLUE, PCT2)
            for j in range(len(gs)):
                wref = f"$A${rr}"; gref = f"{get_column_letter(2+j)}${r0}"
                tvf = (f"DCF!{DCFREF['NOPn']}*(1+{gref})*(1-{gref}/DCF!{DCFREF['ROIC']})"
                       f"/({wref}-{gref})/(1+{wref})^({n}-0.5)")
                f = (f"=(SUMPRODUCT({fcff_rng},1/(1+{wref})^{t_rng})+{tvf}"
                     f"-{A['NETDEBT']}-{A['MIN']})/{A['SHARES']}")
                _w(ws, rr, 2 + j, f, BLACK, PS)


    # ================= SECTOR SHEETS: RI / DDM / COMPS ====================== #
    methods = res.get("methods", {})

    if "residual_income" in methods:
        m = methods["residual_income"]
        ws = wb.create_sheet("Residual income")
        _title(ws, "Residual income model (banks / insurers)",
               "Value = book value + PV of returns earned above the cost of equity. "
               "Blue cells are inputs.")
        ws.column_dimensions["A"].width = 34
        rn = 4
        bvps0 = (data.latest.total_equity - data.latest.minority_interest) / data.shares_diluted
        _w(ws, rn, 1, f"Book value / share ({cur})"); _w(ws, rn, 2, bvps0, BLUE, PS); BV0=f"$B${rn}"; rn+=1
        roe_hist = [y.net_income/y.total_equity for y in data.years if y.total_equity>0]
        roe0 = sum(roe_hist)/len(roe_hist) if roe_hist else 0.10
        _w(ws, rn, 1, "Normalised ROE"); c=_w(ws, rn, 2, roe0, BLUE, PCT); c.fill=YEL; ROE=f"$B${rn}"; rn+=1
        _w(ws, rn, 1, "Cost of equity (Ke)"); _w(ws, rn, 2, f"={KE}", GREEN, PCT2); KEr=f"$B${rn}"; rn+=1
        payout = (data.dividend_per_share/data.latest.eps_diluted
                  if data.latest.eps_diluted and data.dividend_per_share else 0.3)
        payout = min(max(payout,0.0),1.0)
        _w(ws, rn, 1, "Payout ratio"); _w(ws, rn, 2, payout, BLUE, PCT); PAY=f"$B${rn}"; rn+=1
        _w(ws, rn, 1, "Continuing growth (g)"); _w(ws, rn, 2, 0.03, BLUE, PCT); Gr=f"$B${rn}"; rn+=1
        _w(ws, rn, 1, "Book growth (min((1−payout)×ROE, g+2%))")
        _w(ws, rn, 2, f"=MIN((1-{PAY})*{ROE},{Gr}+0.02)", BLACK, PCT); BG=f"$B${rn}"; rn+=2
        nri = 6
        _hdr_row(ws, rn, ["Year"] + [f"Y{i+1}" for i in range(nri)]); rn+=1
        _w(ws, rn, 1, "Opening BV / share")
        _w(ws, rn, 2, f"={BV0}", BLACK, PS)
        for j in range(1, nri):
            cl = get_column_letter(2+j)
            _w(ws, rn, 2+j, f"={get_column_letter(1+j)}{rn}*(1+{BG})", BLACK, PS)
        BVR=rn; rn+=1
        _w(ws, rn, 1, "Residual income ((ROE−Ke)×BV)")
        for j in range(nri):
            cl = get_column_letter(2+j)
            _w(ws, rn, 2+j, f"=({ROE}-{KEr})*{cl}{BVR}", BLACK, PS)
        RIR=rn; rn+=1
        _w(ws, rn, 1, "PV of RI")
        for j in range(nri):
            cl = get_column_letter(2+j)
            _w(ws, rn, 2+j, f"={cl}{RIR}/(1+{KEr})^{j+1}", BLACK, PS)
        PVR=rn; rn+=2
        _w(ws, rn, 1, "PV of explicit RI")
        _w(ws, rn, 2, f"=SUM(B{PVR}:{get_column_letter(1+nri)}{PVR})", BLACK, PS); PVEX=f"$B${rn}"; rn+=1
        _w(ws, rn, 1, "Continuing value (PV)")
        lastbv = f"{get_column_letter(1+nri)}{BVR}"
        _w(ws, rn, 2, f"=({ROE}-{KEr})*{lastbv}*(1+{BG})/({KEr}-{Gr})/(1+{KEr})^{nri}",  # bv_{n+1}=lastbv*(1+BG); engine-consistent
           BLACK, PS); CV=f"$B${rn}"; rn+=1
        _w(ws, rn, 1, "Value per share", BOLD)
        _w(ws, rn, 2, f"={BV0}+{PVEX}+{CV}", BLACK, PS, bold=True, border=DBL); rn+=1

    if "ddm" in methods and data.dividend_per_share:
        ws = wb.create_sheet("DDM")
        _title(ws, "Multi-stage dividend discount",
               "Dividends fade from the retention-implied growth to terminal g, at Ke.")
        ws.column_dimensions["A"].width = 34
        rn = 4
        _w(ws, rn, 1, f"Current DPS ({cur})"); _w(ws, rn, 2, data.dividend_per_share, BLUE, PS); D0=f"$B${rn}"; rn+=1
        _w(ws, rn, 1, "Near-term dividend growth"); c=_w(ws, rn, 2, 0.06, BLUE, PCT); c.fill=YEL; G1=f"$B${rn}"; rn+=1
        _w(ws, rn, 1, "Terminal growth"); _w(ws, rn, 2, f"={A['TG']}", GREEN, PCT); G2=f"$B${rn}"; rn+=1
        _w(ws, rn, 1, "Cost of equity (Ke)"); _w(ws, rn, 2, f"={KE}", GREEN, PCT2); KEd=f"$B${rn}"; rn+=2
        nd = 6
        _hdr_row(ws, rn, ["Year"] + [f"Y{i+1}" for i in range(nd)]); rn+=1
        _w(ws, rn, 1, "Growth (fade)")
        for j in range(nd):
            _w(ws, rn, 2+j, f"={G1}+({G2}-{G1})*{j}/{nd-1}", BLACK, PCT)
        GR=rn; rn+=1
        _w(ws, rn, 1, "DPS")
        _w(ws, rn, 2, f"={D0}*(1+B{GR})", BLACK, PS)
        for j in range(1, nd):
            cl, pv = get_column_letter(2+j), get_column_letter(1+j)
            _w(ws, rn, 2+j, f"={pv}{rn}*(1+{cl}{GR})", BLACK, PS)
        DPS=rn; rn+=1
        _w(ws, rn, 1, "PV of DPS")
        for j in range(nd):
            cl = get_column_letter(2+j)
            _w(ws, rn, 2+j, f"={cl}{DPS}/(1+{KEd})^{j+1}", BLACK, PS)
        PVD=rn; rn+=2
        _w(ws, rn, 1, "PV explicit dividends")
        _w(ws, rn, 2, f"=SUM(B{PVD}:{get_column_letter(1+nd)}{PVD})", BLACK, PS); PVX=f"$B${rn}"; rn+=1
        _w(ws, rn, 1, "PV terminal value")
        lastd = f"{get_column_letter(1+nd)}{DPS}"
        _w(ws, rn, 2, f"={lastd}*(1+{G2})/({KEd}-{G2})/(1+{KEd})^{nd}", BLACK, PS); PVT=f"$B${rn}"; rn+=1
        _w(ws, rn, 1, "Value per share", BOLD)
        _w(ws, rn, 2, f"={PVX}+{PVT}", BLACK, PS, bold=True, border=DBL); rn+=1

    comps = res.get("comps")
    if comps and comps.peers:
        ws = wb.create_sheet("Comps")
        _title(ws, "Comparable companies",
               "Peer inputs in blue; medians and implied values are live formulas.")
        ws.column_dimensions["A"].width = 16
        rn = 4
        _hdr_row(ws, rn, ["Peer", "EV/EBITDA", "EV/EBIT", "EV/Sales", "P/E"]); rn+=1
        first = rn
        for ppe in comps.peers:
            _w(ws, rn, 1, ppe.ticker)
            for j, v in enumerate([ppe.ev_ebitda, ppe.ev_ebit, ppe.ev_sales, ppe.pe]):
                if v: _w(ws, rn, 2+j, v, BLUE, MULT)
            rn += 1
        last = rn - 1
        _w(ws, rn, 1, "Median", BOLD)
        for j in range(4):
            cl = get_column_letter(2+j)
            _w(ws, rn, 2+j, f"=MEDIAN({cl}{first}:{cl}{last})", BLACK, MULT, bold=True, border=TOPLINE)
        MED=rn; rn+=2
        y = data.latest
        metrics = [("EBITDA", (y.ebitda or (y.ebit+y.depreciation_amort))/scale, "B", True),
                   ("EBIT", y.ebit/scale, "C", True),
                   ("Revenue", y.revenue/scale, "D", True),
                   ("EPS", y.eps_diluted, "E", False)]
        _hdr_row(ws, rn, ["Company metric", "Value", "×median", f"Implied VPS ({cur})"]); rn+=1
        for lab, val, mcol_, is_ev in metrics:
            _w(ws, rn, 1, lab)
            _w(ws, rn, 2, val, BLUE, MM if is_ev else PS)
            _w(ws, rn, 3, f"={mcol_}{MED}", BLACK, MULT)
            if is_ev:
                _w(ws, rn, 4, f"=(B{rn}*C{rn}-{A['NETDEBT']}-{A['MIN']})/{A['SHARES']}", GREEN, PS)
            else:
                _w(ws, rn, 4, f"=B{rn}*C{rn}", BLACK, PS)
            rn += 1

    # ===================== REPORTED FINANCIALS ============================== #
    stmts = getattr(data, "raw_statements", None) or {}
    names = {"income": "IS (reported)", "balance": "BS (reported)", "cashflow": "CF (reported)"}
    for key, sheetname in names.items():
        rows = stmts.get(key)
        if not rows:
            continue
        ws = wb.create_sheet(sheetname)
        years = sorted({yr for _, series in rows for yr in series})
        _title(ws, f"{data.ticker} — {sheetname}",
               f"As reported by {data.provider} ({cur}mm). Hardcoded actuals.")
        ws.column_dimensions["A"].width = 42
        r = 4
        _hdr_row(ws, r, ["Line item"] + [f"FY{y}" for y in years]); r += 1
        for label, series in rows:
            _w(ws, r, 1, label)
            for j, y in enumerate(years):
                v = series.get(y)
                if v is not None:
                    _w(ws, r, 2 + j, v / scale, BLUE, MM, fill=HIST_FILL)
            r += 1
        for j in range(len(years)):
            ws.column_dimensions[get_column_letter(2 + j)].width = 13

    # ============================ COVER ===================================== #
    ws = wb.active
    ws.title = "Cover"
    _title(ws, f"{data.name} ({data.ticker})",
           f"{data.sector} · {res['spec'].label} · generated {date.today().isoformat()} · data: {data.provider}")
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 16
    r = 4
    _w(ws, r, 1, "Current price"); _w(ws, r, 2, f"={A['PRICE']}", GREEN, PS); r += 1
    if bd and dcf and MODEL:
        _w(ws, r, 1, "DCF value per share"); _w(ws, r, 2, f"=DCF!$B${DCFREF['VPSROW']}", GREEN, PS); r += 1
        _w(ws, r, 1, "Upside / downside")
        _w(ws, r, 2, f"=DCF!$B${DCFREF['VPSROW']}/{A['PRICE']}-1", GREEN, PCT, bold=True); r += 1
    _w(ws, r, 1, "Blended fair value (report)"); _w(ws, r, 2, res["target"], BLUE, PS); r += 1
    _w(ws, r, 1, "WACC"); _w(ws, r, 2, f"={WACCC}", GREEN, PCT2); r += 2
    _w(ws, r, 1, "How to use", H2); r += 1
    for line in [
        "Blue cells are inputs — edit growth, margins, beta, ERP or terminal g",
        "in the Assumptions sheet and everything recalculates live.",
        "Black cells are formulas; green cells link across sheets.",
        "The Sensitivity grid re-runs the full DCF per WACC/g pair.",
        "Not investment advice. Verify source data before relying on outputs.",
    ]:
        _w(ws, r, 1, line, MUT); r += 1

    wb.save(path)
    return path
